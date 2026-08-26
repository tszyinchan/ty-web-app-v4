# Test-load AppSheet xlsx into tyapp_yyems_* (not the future cutover migration).
#
# Prereqs:
#   1. Paste yyems.schema.sql in Supabase (once).
#   2. tyapp_user.appsheet_525_user_id is 'cty' and 'frd' for the two logins.
#   3. pip install openpyxl supabase
#   4. Env: SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY
#      (Dashboard → Settings → API → service_role; never commit this key.)
#
# Usage:
#   python src/app/apps/jaxfr/features/yyems/import_yyems_xlsx.py --dry-run
#   python src/app/apps/jaxfr/features/yyems/import_yyems_xlsx.py --wipe --xlsx "D:\\path\\backup.xlsx"
#
# --wipe runs the TRUNCATE in yyems.wipe-test-data.sql first (via PostgREST
# deletes). If wipe fails, paste that SQL file in the editor then re-run.
#
# Later: wipe + drop/recreate schema + a separate cutover importer.

from __future__ import annotations

import argparse
import os
import sys
import uuid
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

DEFAULT_XLSX = (
    r"j:\.shortcut-targets-by-id\1yqrgKWWM13JcXnN0VdIaLbLVzW7LytpH"
    r"\6FRD\525\Appsheet data backup\20260825_000400_Items.xlsx"
)

EMAIL_TO_PARTY = {
    "tszyinchan99@gmail.com": "cty",
    "fredchuny1@gmail.com": "frd",
}

MEALS = {"1早", "2茶", "3午", "4茶", "5晚", "6宵", "7用", "8調", "9洗"}
BATCH = 200


def nonempty(v: object) -> bool:
    return v is not None and str(v).strip() != ""


def as_text(v: object) -> str | None:
    if not nonempty(v):
        return None
    return str(v).strip()


def as_bool(v: object) -> bool | None:
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    if s in {"true", "1", "y", "有機", "有味精"}:
        return True
    if s in {"false", "0", "n", "無味精"}:
        return False
    return None


def as_num(v: object) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def as_int(v: object, default: int | None = None) -> int | None:
    n = as_num(v)
    if n is None:
        return default
    return int(n)


def as_iso_dt(v: object) -> str | None:
    if isinstance(v, datetime):
        return v.isoformat()
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day).isoformat()
    if not nonempty(v):
        return None
    s = str(v).strip()
    try:
        return datetime.fromisoformat(s[:19]).isoformat()
    except ValueError:
        return None


def as_iso_date(v: object) -> str | None:
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if not nonempty(v):
        return None
    s = str(v).strip()[:10]
    if len(s) >= 10 and s[4] == "-":
        return s[:10]
    try:
        return datetime.fromisoformat(str(v)[:19]).date().isoformat()
    except ValueError:
        return None


def sheet_rows(wb, name: str) -> list[dict[str, object]]:
    ws = wb[name]
    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(rows)]
    out: list[dict[str, object]] = []
    for row in rows:
        rec: dict[str, object] = {}
        empty = True
        for i, key in enumerate(header):
            if not key:
                continue
            val = row[i] if i < len(row) else None
            rec[key] = val
            if nonempty(val):
                empty = False
        if not empty:
            out.append(rec)
    return out


def new_id() -> str:
    return str(uuid.uuid4())


def chunked(rows: list[dict[str, Any]], size: int = BATCH):
    for i in range(0, len(rows), size):
        yield rows[i : i + size]


class Skip(Exception):
    pass


def main() -> int:
    parser = argparse.ArgumentParser(description="Test-import 525 xlsx into Supabase")
    parser.add_argument("--xlsx", default=DEFAULT_XLSX)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--wipe", action="store_true", help="Delete existing yyems rows first")
    args = parser.parse_args()

    xlsx = Path(args.xlsx)
    if not xlsx.is_file():
        print(f"xlsx not found: {xlsx}", file=sys.stderr)
        return 1

    url = os.environ.get("SUPABASE_URL", "").strip()
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "").strip()
    if not args.dry_run and (not url or not key):
        print("Set SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY", file=sys.stderr)
        return 1

    print(f"Reading {xlsx} …")
    wb = load_workbook(xlsx, read_only=True, data_only=True)

    categories = sheet_rows(wb, "Categories")
    items = sheet_rows(wb, "Items")
    vendor_cats = sheet_rows(wb, "Vendor Categories")
    vendors = sheet_rows(wb, "Vendors")
    fas = sheet_rows(wb, "Financial_Accounts")
    wallets = sheet_rows(wb, "Wallet")
    fx_rows = sheet_rows(wb, "Currency_conversion")
    bills = sheet_rows(wb, "YYEMS")
    prices = sheet_rows(wb, "Prices")
    buys = sheet_rows(wb, "Buy")
    eats = sheet_rows(wb, "Eat")
    files = sheet_rows(wb, "YYEMS_Images")
    wb.close()

    print(
        "rows:",
        f"cat={len(categories)} item={len(items)} vcat={len(vendor_cats)}",
        f"vendor={len(vendors)} fa={len(fas)} wallet={len(wallets)}",
        f"bill={len(bills)} price={len(prices)} buy={len(buys)}",
        f"eat={len(eats)} file={len(files)}",
    )

    client = None
    party_user: dict[str, str] = {}
    default_created_by = ""

    if args.dry_run:
        party_user = {
            "cty": "00000000-0000-0000-0000-000000000001",
            "frd": "00000000-0000-0000-0000-000000000002",
        }
        default_created_by = party_user["cty"]
        print("dry-run: dummy user ids for mapping only")
    else:
        try:
            from supabase import create_client
        except ImportError:
            print("pip install supabase", file=sys.stderr)
            return 1
        client = create_client(url, key)
        users = (
            client.table("tyapp_user")
            .select("user_id, appsheet_525_user_id")
            .is_("deleted_at", "null")
            .execute()
            .data
            or []
        )
        for u in users:
            code = (u.get("appsheet_525_user_id") or "").strip().lower()
            if code in {"cty", "frd"}:
                party_user[code] = u["user_id"]
        if "cty" not in party_user or "frd" not in party_user:
            print(
                "Need tyapp_user.appsheet_525_user_id = cty and frd",
                file=sys.stderr,
            )
            return 1
        default_created_by = party_user["cty"]
        print("mapped users", party_user)

        if args.wipe:
            wipe_tables = [
                "tyapp_yyems_eat",
                "tyapp_yyems_file",
                "tyapp_yyems_buy",
                "tyapp_yyems_price",
                "tyapp_yyems",
                "tyapp_yyems_wallet",
                "tyapp_yyems_financial_account",
                "tyapp_yyems_fx_rate",
                "tyapp_yyems_vendor",
                "tyapp_yyems_vendor_category",
                "tyapp_yyems_item",
                "tyapp_yyems_item_category",
            ]
            for table in wipe_tables:
                client.table(table).delete().gte(
                    "created_at", "1970-01-01T00:00:00Z"
                ).execute()
            print("wiped existing yyems rows")

    def created_by_from_email(email: object) -> str:
        party = EMAIL_TO_PARTY.get((as_text(email) or "").lower(), "")
        return party_user.get(party, default_created_by)

    def person_user(code: object) -> str | None:
        c = (as_text(code) or "").strip().lower()
        if c in {"cty", "frd"}:
            return party_user.get(c)
        return None

    skipped: dict[str, int] = {}

    def skip(kind: str) -> None:
        skipped[kind] = skipped.get(kind, 0) + 1

    cat_ids: dict[str, str] = {}
    item_ids: dict[str, str] = {}
    vcat_ids: dict[str, str] = {}
    vendor_ids: dict[str, str] = {}
    fa_ids: dict[str, str] = {}
    wallet_ids: dict[str, str] = {}
    bill_ids: dict[str, str] = {}
    price_ids: dict[str, str] = {}
    buy_ids: dict[str, str] = {}

    cat_payload: list[dict[str, Any]] = []
    for i, row in enumerate(categories):
        lid = as_text(row.get("ID"))
        if not lid:
            skip("category")
            continue
        pk = new_id()
        cat_ids[lid] = pk
        cat_payload.append(
            {
                "tb_tyapp_yic_id": pk,
                "legacy_id": lid,
                "code": lid,
                "name_zh": as_text(row.get("Category")) or lid,
                "name_en": as_text(row.get("Category_EN")),
                "division": as_text(row.get("Division")),
                "sort_order": i,
            }
        )

    item_payload: list[dict[str, Any]] = []
    for row in items:
        lid = as_text(row.get("Item ID"))
        cat = as_text(row.get("Category"))
        if not lid or not cat or cat not in cat_ids:
            skip("item")
            continue
        pk = new_id()
        item_ids[lid] = pk
        plan = as_bool(row.get("Plan Buy?"))
        item_payload.append(
            {
                "tb_tyapp_yit_id": pk,
                "legacy_id": lid,
                "category_id": cat_ids[cat],
                "name_zh": as_text(row.get("Name")) or lid,
                "name_en": as_text(row.get("Name_EN"))
                or as_text(row.get("Name_EN_auto")),
                "food_category": as_text(row.get("Food Category")),
                "description": as_text(row.get("Description")),
                "plan_buy": bool(plan) if plan is not None else False,
            }
        )

    vcat_payload: list[dict[str, Any]] = []
    for row in vendor_cats:
        lid = as_text(row.get("ID"))
        if not lid:
            skip("vendor_category")
            continue
        pk = new_id()
        vcat_ids[lid] = pk
        vcat_payload.append(
            {
                "tb_tyapp_yvc_id": pk,
                "legacy_id": lid,
                "level1": as_text(row.get("一級分類")) or "",
                "level2": as_text(row.get("二級分類")) or "",
                "level3": as_text(row.get("三級分類")) or "",
                "display_name": as_text(row.get("顯示名稱")) or lid,
            }
        )

    vendor_payload: list[dict[str, Any]] = []
    for row in vendors:
        lid = as_text(row.get("ID"))
        cat = as_text(row.get("Vendor Categories"))
        if not lid or not cat or cat not in vcat_ids:
            skip("vendor")
            continue
        pk = new_id()
        vendor_ids[lid] = pk
        vendor_payload.append(
            {
                "tb_tyapp_yvd_id": pk,
                "legacy_id": lid,
                "category_id": vcat_ids[cat],
                "name": as_text(row.get("Name")) or lid,
                "name_short": as_text(row.get("Name_簡稱_如有")),
                "sort_order": as_int(row.get("Order")),
            }
        )

    fa_payload: list[dict[str, Any]] = []
    for row in fas:
        lid = as_text(row.get("ID"))
        cur = as_text(row.get("Currency")) or "CAD"
        if not lid:
            skip("fa")
            continue
        pk = new_id()
        fa_ids[lid] = pk
        person = (as_text(row.get("Person ID")) or "").lower()
        fa_payload.append(
            {
                "tb_tyapp_yfa_id": pk,
                "legacy_id": lid,
                "owner_user_id": person_user(person),
                "display_name": as_text(row.get("Display Name")) or lid,
                "currency": cur if cur != "FREE" else "FREE",
            }
        )

    wallet_payload: list[dict[str, Any]] = []
    for row in wallets:
        lid = as_text(row.get("ID"))
        fa = as_text(row.get("Financial_Accounts"))
        if not lid or not fa or fa not in fa_ids:
            skip("wallet")
            continue
        pk = new_id()
        wallet_ids[lid] = pk
        wallet_payload.append(
            {
                "tb_tyapp_ywl_id": pk,
                "legacy_id": lid,
                "financial_account_id": fa_ids[fa],
                "name": as_text(row.get("Name")) or lid,
                "sort_order": as_int(row.get("Order")),
                "remarks": as_text(row.get("Remark")),
            }
        )

    fx_payload: list[dict[str, Any]] = []
    seen_fx: set[tuple[str, int]] = set()
    for row in fx_rows:
        cur = as_text(row.get("Currency"))
        year = as_int(row.get("Year"))
        to_cad = as_num(row.get("To CAD"))
        if not cur or year is None or to_cad is None:
            skip("fx")
            continue
        key = (cur, year)
        if key in seen_fx:
            continue
        seen_fx.add(key)
        fx_payload.append(
            {
                "tb_tyapp_yfx_id": new_id(),
                "currency": cur,
                "year": year,
                "to_cad": to_cad,
                "source": as_text(row.get("Source")),
            }
        )

    bill_payload: list[dict[str, Any]] = []
    for row in bills:
        lid = as_text(row.get("YYEMS ID"))
        vendor = as_text(row.get("Vendor ID"))
        wallet = as_text(row.get("Wallet"))
        occurred = as_iso_dt(row.get("DateTime"))
        amount = as_num(row.get("Amount"))
        if not lid or not vendor or vendor not in vendor_ids:
            skip("bill")
            continue
        if not wallet or wallet not in wallet_ids:
            skip("bill_wallet")
            continue
        if not occurred or amount is None:
            skip("bill_required")
            continue
        tz = as_text(row.get("Location TimeZone")) or "TO"
        if tz not in {"HK", "TO"}:
            tz = "TO"
        flow = (as_text(row.get("In_or_out")) or "out").strip().lower()
        if flow not in {"in", "out", "free"}:
            flow = "out"
        cur = as_text(row.get("Currency")) or "CAD"
        own = (as_text(row.get("Ownership")) or "").strip().lower()
        pk = new_id()
        bill_ids[lid] = pk
        tick = as_text(row.get("✔️"))
        bill_payload.append(
            {
                "tb_tyapp_yym_id": pk,
                "legacy_id": lid,
                "occurred_at": occurred,
                "location_tz": tz,
                "in_or_out": flow,
                "vendor_id": vendor_ids[vendor],
                "currency": cur,
                "amount": amount,
                "wallet_id": wallet_ids[wallet],
                "ownership_user_id": person_user(own),
                "remark": as_text(row.get("remark")),
                "description": as_text(row.get("description")),
                "reconciled": bool(tick),
                "wallet_amount": as_num(row.get("Wallet Amount")),
                "period_start": as_iso_date(row.get("start_date")),
                "period_end": as_iso_date(row.get("end_date")),
                "created_by": created_by_from_email(row.get("email_address")),
            }
        )

    price_payload: list[dict[str, Any]] = []
    for row in prices:
        lid = as_text(row.get("Price ID"))
        item = as_text(row.get("Item ID"))
        priced = as_iso_dt(row.get("Datetime"))
        if not lid or not item or item not in item_ids or not priced:
            skip("price")
            continue
        vendor = as_text(row.get("Vendor ID"))
        btype = as_text(row.get("Barcode_type"))
        bmap = {"UPC": "upc", "PLU": "plu", "Price-embedded": "price_embedded"}
        organic = as_bool(row.get("有機?"))
        msg = as_bool(row.get("有味精?"))
        if as_text(row.get("有味精?")) == "無味精":
            msg = False
        pk = new_id()
        price_ids[lid] = pk
        price_payload.append(
            {
                "tb_tyapp_ypr_id": pk,
                "legacy_id": lid,
                "priced_at": priced,
                "vendor_id": vendor_ids.get(vendor) if vendor else None,
                "item_id": item_ids[item],
                "product_name": as_text(row.get("Product Name")),
                "product_name_zh": as_text(row.get("ProductName_中文")),
                "brand": as_text(row.get("Brand")),
                "currency": as_text(row.get("Currency")) or "CAD",
                "packed_price": as_num(row.get("Packed Price")),
                "tax_rate": as_num(row.get("Tax %")),
                "discount_rate": as_num(row.get("Discount")),
                "marked_price": as_num(row.get("Marked Price")),
                "marked_amount": as_num(row.get("Marked Amount")),
                "marked_unit": as_text(row.get("Marked Unit")),
                "packed_amount": as_num(row.get("Packed Amount")),
                "packed_unit": as_text(row.get("Packed Unit")),
                "tag": as_text(row.get("Tag")),
                "is_organic": organic,
                "has_msg": msg,
                "origin": as_text(row.get("原產地")),
                "barcode": as_text(row.get("Barcode")),
                "barcode_type": bmap.get(btype or "", None),
                "remarks": as_text(row.get("Remarks")),
                "nutri_basis_amount": as_num(row.get("Nutri_Basis_Amount")),
                "nutri_basis_unit": as_text(row.get("Nutri_Basis_Unit")),
                "protein_g": as_num(row.get("Protein_g")),
                "carb_g": as_num(row.get("Carb_g")),
                "fat_g": as_num(row.get("Fat_g")),
                "calories_kcal": as_num(row.get("Calories_kcal")),
                "fiber_g": as_num(row.get("Fiber_g")),
                "sodium_mg": as_num(row.get("Sodium_mg")),
                "nutri_is_estimated": bool(as_bool(row.get("Nutri_Is_Estimated")) or False),
                "created_by": default_created_by,
            }
        )

    buy_payload: list[dict[str, Any]] = []
    for row in buys:
        lid = as_text(row.get("Buy ID"))
        price = as_text(row.get("Price Log ID"))
        home = as_num(row.get("Home Amount"))
        if not lid or not price or price not in price_ids or home is None:
            skip("buy")
            continue
        bill = as_text(row.get("YYEMS ID"))
        pk = new_id()
        buy_ids[lid] = pk
        # PostgREST bulk insert unions keys across the batch; a missing
        # created_at on one row would send NULL for the whole batch and
        # override DEFAULT now(). Always send a timestamp.
        created = as_iso_dt(row.get("Created at")) or datetime.now(
            timezone.utc
        ).isoformat()
        buy_payload.append(
            {
                "tb_tyapp_yby_id": pk,
                "legacy_id": lid,
                "price_id": price_ids[price],
                "yyems_id": bill_ids.get(bill) if bill else None,
                "paid": as_num(row.get("Paid")),
                "home_amount": home,
                "home_unit": as_text(row.get("Home Unit")),
                "marked_amount_count": as_num(row.get("Number of Marked Amounts")),
                "expiry_date": as_iso_date(row.get("Expiry Date")),
                "remarks": as_text(row.get("Remarks")),
                "paid_adjust_note": as_text(row.get("調整")),
                "paid_adjust_reason": as_text(row.get("調整原因")),
                "eat_priority": as_int(row.get("Priority_in_Eat"), 50) or 50,
                "created_by": default_created_by,
                "created_at": created,
            }
        )

    eat_payload: list[dict[str, Any]] = []
    for row in eats:
        lid = as_text(row.get("Eat ID"))
        buy = as_text(row.get("Buy ID"))
        home = as_num(row.get("Home Amount"))
        meal = as_text(row.get("Meal"))
        eat_date = as_iso_date(row.get("Eat Date"))
        if not lid or not buy or buy not in buy_ids or home is None:
            skip("eat")
            continue
        if meal not in MEALS:
            skip("eat_meal")
            continue
        if not eat_date:
            skip("eat_date")
            continue
        who = (as_text(row.get("Eaten By")) or "").strip().lower()
        eaten_user = person_user(who)
        eaten_other = None
        if who in {"yyems"}:
            eaten_other = "shared"
        elif who in {"dining_out", "外"}:
            eaten_other = "dining_out"
        elif who in {"nil", "無"}:
            eaten_other = "nil"
        elif who and not eaten_user:
            skip("eat_who")
            continue
        added = as_iso_dt(row.get("Add Datetime")) or eat_date
        eat_payload.append(
            {
                "tb_tyapp_yet_id": new_id(),
                "legacy_id": lid,
                "buy_id": buy_ids[buy],
                "home_amount": home,
                "meal": meal,
                "eaten_by_user_id": eaten_user,
                "eaten_by_other": eaten_other,
                "eat_date": eat_date,
                "added_at": added,
                "description": as_text(row.get("Description")),
                "created_by": default_created_by,
            }
        )

    file_payload: list[dict[str, Any]] = []
    for row in files:
        lid = as_text(row.get("YYEMS Image ID"))
        bill = as_text(row.get("YYEMS ID"))
        if not lid or not bill or bill not in bill_ids:
            skip("file")
            continue
        kind_raw = (as_text(row.get("Type of Image")) or "photo").lower()
        kind = "receipt" if kind_raw == "receipt" else "photo"
        path = as_text(row.get("Image")) or as_text(row.get("File"))
        if as_text(row.get("File")) and not as_text(row.get("Image")):
            kind = "file"
        if not path:
            skip("file_path")
            continue
        file_payload.append(
            {
                "tb_tyapp_yfl_id": new_id(),
                "legacy_id": lid,
                "yyems_id": bill_ids[bill],
                "kind": kind,
                "drive_file_id": None,
                "legacy_path": path,
                "original_filename": path.split("/")[-1],
                "created_by": default_created_by,
            }
        )

    print("payloads:", {k: len(v) for k, v in {
        "cat": cat_payload,
        "item": item_payload,
        "vcat": vcat_payload,
        "vendor": vendor_payload,
        "fa": fa_payload,
        "wallet": wallet_payload,
        "fx": fx_payload,
        "bill": bill_payload,
        "price": price_payload,
        "buy": buy_payload,
        "eat": eat_payload,
        "file": file_payload,
    }.items()})
    print("skipped:", skipped)

    if args.dry_run or client is None:
        print("dry-run: no writes")
        return 0

    def insert_all(table: str, rows: list[dict[str, Any]]) -> None:
        if not rows:
            return
        for part in chunked(rows):
            clean = [{k: v for k, v in row.items() if v is not None} for row in part]
            client.table(table).insert(clean).execute()
        print(f"inserted {table} {len(rows)}")

    insert_all("tyapp_yyems_item_category", cat_payload)
    insert_all("tyapp_yyems_item", item_payload)
    insert_all("tyapp_yyems_vendor_category", vcat_payload)
    insert_all("tyapp_yyems_vendor", vendor_payload)
    insert_all("tyapp_yyems_financial_account", fa_payload)
    insert_all("tyapp_yyems_wallet", wallet_payload)
    insert_all("tyapp_yyems_fx_rate", fx_payload)
    insert_all("tyapp_yyems", bill_payload)
    insert_all("tyapp_yyems_price", price_payload)
    insert_all("tyapp_yyems_buy", buy_payload)
    insert_all("tyapp_yyems_eat", eat_payload)
    insert_all("tyapp_yyems_file", file_payload)
    print("done")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
